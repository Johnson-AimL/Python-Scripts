# this is the current working code for excel sort and report
# for electrical company.


import os
import wx
import pandas as pd
import openpyxl
import xlrd
import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font


class mainApp(wx.Frame):

    def __init__(self, *args, **kwargs):
        super(mainApp, self).__init__(*args, **kwargs)
        self.InitUI()

    def InitUI(self):
        self.SetBackgroundColour(wx.BLUE)
        menubar = wx.MenuBar()
        fileMenu = wx.Menu()
        fileItem = fileMenu.Append(wx.ID_EXIT, 'Quit', 'Quit application')
        menubar.Append(fileMenu, '&File')
        self.SetMenuBar(menubar)
        self.Bind(wx.EVT_MENU, self.OnQuit, fileItem)

        font = wx.Font(pointSize=16, family=wx.FONTFAMILY_DEFAULT, style=wx.FONTSTYLE_NORMAL,
                       weight=wx.FONTWEIGHT_NORMAL)

        vbox = wx.BoxSizer(wx.VERTICAL)

        hbox1 = wx.BoxSizer(wx.HORIZONTAL)
        stock_label = wx.StaticText(self, label='Select Stock Excel:')
        stock_label.SetFont(font)
        hbox1.Add(stock_label, flag=wx.LEFT | wx.RIGHT, border=10)
        vbox.Add(hbox1, flag=wx.LEFT | wx.RIGHT | wx.TOP, border=10)

        hbox2 = wx.BoxSizer(wx.HORIZONTAL)
        self.file1 = wx.FilePickerCtrl(self, style=wx.FLP_USE_TEXTCTRL)
        self.file1.SetFont(font)
        hbox2.Add(self.file1, proportion=1, flag=wx.EXPAND)
        vbox.Add(hbox2, flag=wx.LEFT | wx.RIGHT | wx.TOP, border=10)

        vbox.Add((-1, 20))

        hbox3 = wx.BoxSizer(wx.HORIZONTAL)
        tally_label = wx.StaticText(self, label='Select Tally Excel:')
        tally_label.SetFont(font)
        hbox3.Add(tally_label, flag=wx.LEFT | wx.RIGHT, border=10)
        vbox.Add(hbox3, flag=wx.LEFT | wx.RIGHT | wx.TOP, border=10)

        hbox4 = wx.BoxSizer(wx.HORIZONTAL)
        self.file2 = wx.FilePickerCtrl(self, style=wx.FLP_USE_TEXTCTRL)
        hbox4.Add(self.file2, proportion=1, flag=wx.EXPAND)
        vbox.Add(hbox4, flag=wx.LEFT | wx.RIGHT | wx.TOP, border=10)

        vbox.Add((-1, 20))

        hbox5 = wx.BoxSizer(wx.HORIZONTAL)
        vyapar_label = wx.StaticText(self, label='Select Vyapar Excel:')
        vyapar_label.SetFont(font)
        hbox5.Add(vyapar_label, flag=wx.LEFT | wx.RIGHT, border=10)
        vbox.Add(hbox5, flag=wx.LEFT | wx.RIGHT | wx.TOP, border=10)

        hbox6 = wx.BoxSizer(wx.HORIZONTAL)
        self.file3 = wx.FilePickerCtrl(self, style=wx.FLP_USE_TEXTCTRL)
        hbox6.Add(self.file3, proportion=1, flag=wx.EXPAND)
        vbox.Add(hbox6, flag=wx.LEFT | wx.RIGHT | wx.TOP, border=10)

        vbox.Add((-1, 20))

        hbox7 = wx.BoxSizer(wx.HORIZONTAL)
        submit_button = wx.Button(self, label='Submit')
        hbox7.Add(submit_button, proportion=1, flag=wx.EXPAND)
        vbox.Add(hbox7, flag=wx.LEFT | wx.BOTTOM, border=10)

        self.Bind(wx.EVT_BUTTON, self.on_submit, submit_button)

        vbox.Add((-1, 20))

        font = wx.Font(10, wx.DECORATIVE, wx.ITALIC, wx.NORMAL)

        hbox8 = wx.BoxSizer(wx.HORIZONTAL)
        self.label = wx.StaticText(self, label="Welcome, RJW ELECTRICALS  - Stock Matching Program V1.0")
        self.label.SetMinSize((400, 30))  # set the minimum size of the text box
        self.label.SetFont(font)
        #     self.label.SetBackgroundColour("#FFFFFF")

        hbox8.Add(self.label, proportion=1, flag=wx.EXPAND)
        vbox.Add(hbox8, flag=wx.LEFT | wx.BOTTOM, border=10)

        # vbox.AddSpacer(100)  # Add some padding to the top

        self.gauge = wx.Gauge(self, range=100, size=(-1, 25), style=wx.GA_HORIZONTAL)
        self.gauge.SetForegroundColour(wx.Colour(0, 255, 0))  # Green color
        vbox.Add(self.gauge, proportion=0, flag=wx.EXPAND)

        self.SetSizer(vbox)

        self.update_progress(0)
        #  self.count = 0
        #  self.timer = wx.Timer(self)
        #  self.Bind(wx.EVT_TIMER, self.OnTimer)
        #  self.timer.Start(100)

        self.SetSize((700, 500))
        self.SetTitle('RJW ELECTRICALS  - Stock Matching Program V1.0')
        self.Centre()

        self.Show()

    def loadVyapar(self):

        self.label.SetLabel("Processing Vyapar excel sheet, Please wait.")
        self.Layout()

        # Load the XLS file into a variable
        workbookT = xlrd.open_workbook(self.file3.GetPath())
        sheetV = workbookT.sheet_by_index(0)

        # Create an empty DataFrame to store the item and quantity data
        self.vyaparitem_df = pd.DataFrame(columns=["Item", "Quantity"])

        # Loop through each row in the sheet
        for i in range(sheetV.nrows):
            row = sheetV.row_values(i)
            # Check if both the item and quantity have a value
            #         if (row[1] and row[2] and row[1] != "Item Name") or (row[1] and row[2]== 0):
            if (row[0] and row[1] and row[0] != "Item Name"):
                # Add the item and quantity to the DataFrame
                self.vyaparitem_df = self.vyaparitem_df.append({"Item": row[0], "Quantity": row[1]}, ignore_index=True)

        # Changing the datatype of item to string for sorting
        self.vyaparitem_df['Item'] = self.vyaparitem_df['Item'].astype(str)

        # Sort the dataframe based on 'Item' column
        self.vyaparitem_df.sort_values('Item', inplace=True)

        # Remove leading and trailing spaces from string columns
        self.vyaparitem_df['Item'] = self.vyaparitem_df['Item'].astype(str).str.strip()

        # replacing Quantity field as int and no decimal place
        self.vyaparitem_df['Quantity'] = self.vyaparitem_df['Quantity'].astype(str).str.replace('.', '',
                                                                                                regex=False).astype(int)

        self.vyaparitem_dfU = self.vyaparitem_df.groupby('Item', as_index=False).sum()

        self.vyaparitem_dfD = self.vyaparitem_df[self.vyaparitem_df.duplicated(subset='Item', keep=False)].sort_values(
            'Item')

    def loadTally(self):

        self.label.SetLabel("Processing Tally excel sheet, Please wait.")
        self.Layout()

        # Load the Excel file into a variable
        workbook = openpyxl.load_workbook(self.file2.GetPath())
        sheetT = workbook.active

        # Create an empty DataFrame to store the item and quantity data

        self.tallyitem_df = pd.DataFrame(columns=["Item", "Quantity"])

        # Loop through each row in the sheet
        row_num = 1
        col_num = 1
        for row in sheetT.iter_rows(values_only=True):
            # Check if the font style of cell A is italic
            if sheetT.cell(row=row_num, column=col_num).font.italic:

                # Check if the quantity is None or spaces
                if row[1] is None or row[1] == "":
                    quantity = 0
                else:
                    quantity = row[1]
                # Cell A is italic, add the item and quantity to the DataFrame
                self.tallyitem_df = self.tallyitem_df.append({"Item": row[0], "Quantity": quantity}, ignore_index=True)
            row_num += 1

        # changing the datatype of item to string for sorting
        self.tallyitem_df['Item'] = self.tallyitem_df['Item'].astype(str)

        # sort the dataframe based on 'Item' column
        self.tallyitem_df.sort_values('Item', inplace=True)

        # remove leading and trailing spaces from string columns
        # self.tallyitem_df = self.tallyitem_df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
        self.tallyitem_df['Item'] = self.tallyitem_df['Item'].astype(str).str.strip()

        self.tallyitem_dfU = self.tallyitem_df.groupby('Item', as_index=False).sum()

        self.tallyitem_dfD = self.tallyitem_df[self.tallyitem_df.duplicated(subset='Item', keep=False)].sort_values(
            'Item')

    def loadStock(self):

        self.update_progress(12)
        self.label.SetLabel("Processing Stock excel sheet, Please wait.")

        # Load the Excel file into a variable
        workbook = openpyxl.load_workbook(self.file1.GetPath())
        sheetS = workbook.active

        # Create an empty DataFrame to store the item and quantity data

        self.Stockitem_df = pd.DataFrame(columns=["Item", "Quantity"])

        self.update_progress(14)

        # Loop through each row in the sheet
        for i, row in enumerate(sheetS.iter_rows(values_only=True)):
            # Check if both the item and quantity have a value
            #       if row[1] and row[2] and row[1] != "PRODUCT REFERENCE":
            if (row[1] and row[2] and row[1] != "PRODUCT REFERENCE") or (row[1] and row[2] == 0):
                # Add the item and quantity to the DataFrame
                self.Stockitem_df = self.Stockitem_df.append({"Item": row[1], "Quantity": row[2]}, ignore_index=True)

        self.update_progress(16)

        # changing the datatype of item to string for sorting
        self.Stockitem_df['Item'] = self.Stockitem_df['Item'].astype(str)

        # sort the dataframe based on 'Item' column
        self.Stockitem_df.sort_values('Item', inplace=True)

        self.update_progress(18)

        # remove leading and trailing spaces from string columns
        #   self.Stockitem_df = self.Stockitem_df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
        self.Stockitem_df['Item'] = self.Stockitem_df['Item'].astype(str).str.strip()

        self.Stockitem_dfU = self.Stockitem_df.groupby('Item', as_index=False).sum()

        self.Stockitem_dfD = self.Stockitem_df[self.Stockitem_df.duplicated(subset='Item', keep=False)].sort_values(
            'Item')

    def getcommonlist(self):

        self.label.SetLabel("Generating matching list, Please wait.")
        self.Layout()

        merged_df = pd.merge(self.tallyitem_dfU, self.Stockitem_dfU, on='Item', how='inner')
        merged_df = pd.merge(merged_df, self.vyaparitem_dfU, on='Item', how='inner')

        self.result = merged_df.groupby('Item').agg({'Quantity_x': 'sum', 'Quantity_y': 'sum', 'Quantity': 'sum'})
        self.result = self.result.reset_index()
        # Rename the columns to indicate the source of information
        self.result = self.result.rename(
            columns={'Quantity_x': 'Tally_Value', 'Quantity_y': 'Stock_Excel_Value', 'Quantity': 'Vyapar_Value'})

        # Group the data by the Item column
        #    grouped = self.result.groupby('Item')

        # Use the agg method to check if all three quantities are equal
        #    equal_quantities = grouped['Tally_Value', 'Stock_Excel_Value', 'Vyapar_Value'].agg(lambda x: x.nunique() == 1)

        self.matching_items = self.result[(self.result['Tally_Value'] == self.result['Stock_Excel_Value']) & (
                    self.result['Tally_Value'] == self.result['Vyapar_Value'])]
        self.matching_items = self.matching_items[['Item', 'Tally_Value', 'Stock_Excel_Value', 'Vyapar_Value']]

        self.not_matching_items = self.result[(self.result['Tally_Value'] != self.result['Stock_Excel_Value']) | (
                    self.result['Tally_Value'] != self.result['Vyapar_Value'])]
        self.not_matching_items = self.not_matching_items[['Item', 'Tally_Value', 'Stock_Excel_Value', 'Vyapar_Value']]

        # Create two separate dataframes for matching and non-matching quantities

    #      self.matching = equal_quantities[equal_quantities == True].reset_index()
    #     self.not_matching = equal_quantities[equal_quantities == False].reset_index()

    # Join the original result dataframe to the matching dataframe
    #     self.matching = self.matching.merge(self.result, on='Item', how='inner')

    # Join the original result dataframe to the not matching dataframe
    #    self.not_matching = self.not_matching.merge(self.result, on='Item', how='inner')

    def missingItems(self):

        self.label.SetLabel("Processing Missing items list, Please wait.")
        self.Layout()

        # Merge the dataframes with outer join
        df_merged = pd.merge(self.tallyitem_df, self.Stockitem_df, on='Item', how='outer')
        df_merged = pd.merge(df_merged, self.vyaparitem_df, on='Item', how='outer')

        # Create a dataframe to keep the information of missing items
        missing_items = df_merged[df_merged.isna().any(axis=1)]
        self.missing_df = pd.DataFrame(columns=['Item', 'Tally', 'Stock_Excel', 'Vyapar'])

        # Find out which dataframe the missing items are in
        for index, row in missing_items.iterrows():
            item = row['Item']
            if item not in self.tallyitem_df['Item'].values:
                self.missing_df = self.missing_df.append({'Item': item, 'Tally': 'Not Available'}, ignore_index=True)
            else:
                tally_qty = self.tallyitem_df[self.tallyitem_df['Item'] == item]['Quantity'].values[0]
                self.missing_df = self.missing_df.append({'Item': item, 'Tally': tally_qty}, ignore_index=True)

            if item not in self.Stockitem_df['Item'].values:
                self.missing_df = self.missing_df.append({'Item': item, 'Stock_Excel': 'Not Available'},
                                                         ignore_index=True)
            else:
                stock_qty = self.Stockitem_df[self.Stockitem_df['Item'] == item]['Quantity'].values[0]
                self.missing_df = self.missing_df.append({'Item': item, 'Stock_Excel': stock_qty}, ignore_index=True)

            if item not in self.vyaparitem_df['Item'].values:
                self.missing_df = self.missing_df.append({'Item': item, 'Vyapar': 'Not Available'}, ignore_index=True)
            else:
                vyapar_qty = self.vyaparitem_df[self.vyaparitem_df['Item'] == item]['Quantity'].values[0]
                self.missing_df = self.missing_df.append({'Item': item, 'Vyapar': vyapar_qty}, ignore_index=True)

        # Rename the columns to indicate the source of information
        self.missing_df = self.missing_df.rename(
            columns={'Tally': 'Tally_Value', 'Stock_Excel': 'Stock_Excel_Value', 'Vyapar': 'Vyapar_Value'})

        # Group the rows by the item
        grouped = self.missing_df.groupby('Item')

        # Merge the information in the respective columns
        self.merged_df = grouped.agg({'Tally_Value': 'first', 'Stock_Excel_Value': 'first', 'Vyapar_Value': 'first'})
        self.merged_df = self.merged_df.reset_index()

    def writeoutput(self):

        self.label.SetLabel("Formatting output sheet, Please wait.")
        self.Layout()

        # Set the header colors
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')

        # Get the current date and time
        now = datetime.datetime.now()

        #     directory = "D:/files_compare/out"
        self.directory = self.directory + '\Results'

        if not os.path.exists(self.directory):
            os.makedirs(self.directory)

        self.file_name = self.directory + '\Output_' + now.strftime("%Y-%m-%d_%H-%M-%S") + '.xlsx'

        # Create a Workbook object
        book = Workbook()

        # Create a writer object
        writer = pd.ExcelWriter(self.file_name, engine='openpyxl')
        writer.book = book

        ################################################################################################################################
        #  Writing Tallyitem sheet, formatting it and hiding it in the excel.
        #
        ###############################################################################################################################
        # Write each dataframe to a separate tab in the excel file
        self.tallyitem_df.to_excel(writer, index=False, sheet_name='tally_Item')

        # Get a reference to the worksheet
        tallyItemSh = writer.book['tally_Item']

        # Set the column widths
        tallyItemSh.column_dimensions['A'].width = 50
        tallyItemSh.column_dimensions['B'].width = 20

        for col_num, value in enumerate(self.tallyitem_df.columns.values):
            cell = tallyItemSh.cell(row=1, column=col_num + 1)
            cell.value = value
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Iterate through each cell in the worksheet
        for row in tallyItemSh.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                    cell.font = Font(color='FF0000')
                    cell.border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))

        # Hide the worksheet
        tallyItemSh.sheet_state = 'hidden'

        ################################################################################################################################
        #  Writing Stockitem sheet, formatting it and hiding it in the excel.
        #
        ###############################################################################################################################

        self.Stockitem_df.to_excel(writer, index=False, sheet_name='stock_Item')

        # Get a reference to the worksheet
        StockitemSh = writer.book['stock_Item']

        # Set the column widths
        StockitemSh.column_dimensions['A'].width = 50
        StockitemSh.column_dimensions['B'].width = 20

        for col_num, value in enumerate(self.Stockitem_df.columns.values):
            cell = StockitemSh.cell(row=1, column=col_num + 1)
            cell.value = value
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Iterate through each cell in the worksheet
        for row in StockitemSh.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                    cell.font = Font(color='FF0000')
                    cell.border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))

        # Hide the worksheet
        StockitemSh.sheet_state = 'hidden'

        ################################################################################################################################
        #  Writing vyaparitem sheet, formatting it and hiding it in the excel.
        #
        ###############################################################################################################################

        self.vyaparitem_df.to_excel(writer, index=False, sheet_name='Vypar_Item')

        # Get a reference to the worksheet
        vyaparitemSh = writer.book['Vypar_Item']

        # Set the column widths
        vyaparitemSh.column_dimensions['A'].width = 50
        vyaparitemSh.column_dimensions['B'].width = 20

        for col_num, value in enumerate(self.vyaparitem_df.columns.values):
            cell = vyaparitemSh.cell(row=1, column=col_num + 1)
            cell.value = value
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Iterate through each cell in the worksheet
        for row in vyaparitemSh.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                    cell.font = Font(color='FF0000')
                    cell.border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))

        # Hide the worksheet
        vyaparitemSh.sheet_state = 'hidden'

        self.getcommonlist()

        ################################################################################################################################
        #  Writing uniqueItems sheet of three excel, formatting it and hiding it in the excel.
        #
        ###############################################################################################################################

        self.result.to_excel(writer, index=False, sheet_name='ITMS_AVLBL_IN_ALL')

        # Get the worksheet object
        unique_dfSh = writer.sheets['ITMS_AVLBL_IN_ALL']

        # Set the column widths
        unique_dfSh.column_dimensions['A'].width = 50
        unique_dfSh.column_dimensions['B'].width = 20
        unique_dfSh.column_dimensions['C'].width = 20
        unique_dfSh.column_dimensions['D'].width = 20

        for col_num, value in enumerate(self.result.columns.values):
            cell = unique_dfSh.cell(row=1, column=col_num + 1)
            cell.value = value
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Iterate through each cell in the worksheet
        for row in unique_dfSh.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                    cell.font = Font(color='FF0000')
                    cell.border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))

        unique_dfSh.sheet_state = 'hidden'

        ################################################################################################################################
        #  Processing missing items details.
        #
        ###############################################################################################################################

        self.missingItems()

        ################################################################################################################################
        #  Writing uniqueItems sheet of three excel, formatting it and hiding it in the excel.
        #
        ###############################################################################################################################

        #       self.missing_df.to_excel(writer, index=False, sheet_name='missingItems')
        self.merged_df.to_excel(writer, index=False, sheet_name='ITMS_Missing')

        # Get the worksheet object
        merged_dfSh = writer.sheets['ITMS_Missing']

        # Set the column widths
        merged_dfSh.column_dimensions['A'].width = 50
        merged_dfSh.column_dimensions['B'].width = 20
        merged_dfSh.column_dimensions['C'].width = 20
        merged_dfSh.column_dimensions['D'].width = 20

        for col_num, value in enumerate(self.missing_df.columns.values):
            cell = merged_dfSh.cell(row=1, column=col_num + 1)
            cell.value = value
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Iterate through each cell in the worksheet
        for row in merged_dfSh.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                    cell.font = Font(color='FF0000')
                    cell.border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))

        ################################################################################################################################
        #  Writing Tally Duplicate if it exist in sheet , formatting it in the excel.
        #
        ###############################################################################################################################

        if not self.tallyitem_dfD.empty:
            self.tallyitem_dfD.to_excel(writer, index=False, sheet_name='Duplicate_ITM_IN_TLY')

            # Get a reference to the worksheet
            tlydupitemSh = writer.book['Duplicate_ITM_IN_TLY']

            # Set the column widths
            tlydupitemSh.column_dimensions['A'].width = 50
            tlydupitemSh.column_dimensions['B'].width = 20

            for col_num, value in enumerate(self.tallyitem_dfD.columns.values):
                cell = tlydupitemSh.cell(row=1, column=col_num + 1)
                cell.value = value
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # Iterate through each cell in the worksheet
            for row in tlydupitemSh.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)) and cell.value < 0:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                        cell.font = Font(color='FF0000')
                        cell.border = Border(left=Side(style='thin'),
                                             right=Side(style='thin'),
                                             top=Side(style='thin'),
                                             bottom=Side(style='thin'))
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = Border(left=Side(style='thin'),
                                             right=Side(style='thin'),
                                             top=Side(style='thin'),
                                             bottom=Side(style='thin'))

        ################################################################################################################################
        #  Writing Stock Duplicate if it exist in sheet , formatting it in the excel.
        #
        ###############################################################################################################################

        if not self.Stockitem_dfD.empty:
            self.Stockitem_dfD.to_excel(writer, index=False, sheet_name='Duplicate_ITM_IN_STCKXL')

            # Get a reference to the worksheet
            stkdupitemSh = writer.book['Duplicate_ITM_IN_STCKXL']

            # Set the column widths
            stkdupitemSh.column_dimensions['A'].width = 50
            stkdupitemSh.column_dimensions['B'].width = 20

            for col_num, value in enumerate(self.Stockitem_dfD.columns.values):
                cell = stkdupitemSh.cell(row=1, column=col_num + 1)
                cell.value = value
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # Iterate through each cell in the worksheet
            for row in stkdupitemSh.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)) and cell.value < 0:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                        cell.font = Font(color='FF0000')
                        cell.border = Border(left=Side(style='thin'),
                                             right=Side(style='thin'),
                                             top=Side(style='thin'),
                                             bottom=Side(style='thin'))
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = Border(left=Side(style='thin'),
                                             right=Side(style='thin'),
                                             top=Side(style='thin'),
                                             bottom=Side(style='thin'))

        ################################################################################################################################
        #  Writing Vyapar Duplicate if it exist in sheet , formatting it in the excel.
        #
        ###############################################################################################################################

        if not self.vyaparitem_dfD.empty:
            self.vyaparitem_dfD.to_excel(writer, index=False, sheet_name='Duplicate_ITM_IN_VYPR')

            # Get a reference to the worksheet
            vprdupitemSh = writer.book['Duplicate_ITM_IN_VYPR']

            # Set the column widths
            vprdupitemSh.column_dimensions['A'].width = 50
            vprdupitemSh.column_dimensions['B'].width = 20

            for col_num, value in enumerate(self.vyaparitem_dfD.columns.values):
                cell = vprdupitemSh.cell(row=1, column=col_num + 1)
                cell.value = value
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # Iterate through each cell in the worksheet
            for row in vprdupitemSh.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)) and cell.value < 0:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                        cell.font = Font(color='FF0000')
                        cell.border = Border(left=Side(style='thin'),
                                             right=Side(style='thin'),
                                             top=Side(style='thin'),
                                             bottom=Side(style='thin'))
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = Border(left=Side(style='thin'),
                                             right=Side(style='thin'),
                                             top=Side(style='thin'),
                                             bottom=Side(style='thin'))

        ################################################################################################################################
        #  Writing unique items with values matching, formatting it in the excel.
        #
        ###############################################################################################################################

        self.matching_items.to_excel(writer, index=False, sheet_name='Unique_Match')

        # Get a reference to the worksheet
        unqmtchitemSh = writer.book['Unique_Match']

        # Set the column widths
        unqmtchitemSh.column_dimensions['A'].width = 50
        unqmtchitemSh.column_dimensions['B'].width = 20
        unqmtchitemSh.column_dimensions['C'].width = 20
        unqmtchitemSh.column_dimensions['D'].width = 20

        for col_num, value in enumerate(self.matching_items.columns.values):
            cell = unqmtchitemSh.cell(row=1, column=col_num + 1)
            cell.value = value
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Iterate through each cell in the worksheet
        for row in unqmtchitemSh.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                    cell.font = Font(color='FF0000')
                    cell.border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))

        ################################################################################################################################
        #  Writing unique items with values not matching, formatting it in the excel.
        #
        ###############################################################################################################################

        self.not_matching_items.to_excel(writer, index=False, sheet_name='Unique_Mismatch')

        # Get a reference to the worksheet
        unqmismtchitemSh = writer.book['Unique_Mismatch']

        # Set the column widths
        unqmismtchitemSh.column_dimensions['A'].width = 50
        unqmismtchitemSh.column_dimensions['B'].width = 20
        unqmismtchitemSh.column_dimensions['C'].width = 20
        unqmismtchitemSh.column_dimensions['D'].width = 20

        for col_num, value in enumerate(self.not_matching_items.columns.values):
            cell = unqmismtchitemSh.cell(row=1, column=col_num + 1)
            cell.value = value
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Iterate through each cell in the worksheet
        for row in unqmismtchitemSh.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                    cell.font = Font(color='FF0000')
                    cell.border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))

        # Delete the sheet with the specified name
        if 'Sheet' in writer.book.sheetnames:
            writer.book.remove(writer.book['Sheet'])

        # Save the excel file
        writer.save()

        self.label.SetLabel("Sheets updated, Please wait,,,,.")
        self.Layout()

    def OnQuit(self, e):
        self.Close()

    def on_submit(self, event):
        self.label.SetLabel("Processing Started...   Please wait for the update")
        StockFile_path = self.file1.GetPath()
        TallyFile_path = self.file2.GetPath()
        VyaparFile_path = self.file3.GetPath()

        self.directory = os.path.dirname(StockFile_path)
        print("Directory:", self.directory)

        self.loadStock()
        self.update_progress(20)

        self.loadTally()
        self.label.SetLabel("Processing Tally sheet, Please wait.")
        self.update_progress(40)
        self.Layout()

        self.loadVyapar()
        self.label.SetLabel("Processing Vyapar sheet, Please wait.")
        self.update_progress(50)
        self.Layout()
        self.writeoutput()
        self.label.SetLabel("Processing mismatch, Please wait.")
        self.update_progress(90)
        self.Layout()

        print("Button clicked")
        self.update_progress(100)

        self.label.SetLabel("Processing Completed.Results sheet is available in " + self.directory)

    def update_progress(self, value):
        self.gauge.SetValue(value)


#   def OnTimer(self, event):
#      self.count += 1
#     if self.count >= 100:
#        self.timer.Stop()
#    self.gauge.SetValue(self.count)

def main():
    app = wx.App()
    ex = mainApp(None)
    ex.Show()
    app.MainLoop()


if __name__ == '__main__':
    main()
