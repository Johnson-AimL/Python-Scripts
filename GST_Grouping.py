# GST Sorting Program

import os
import wx
import pandas as pd
import openpyxl
import xlrd
import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font, numbers


class mainApp(wx.Frame):

    def __init__(self, *args, **kwargs):
        super(mainApp, self).__init__(*args, **kwargs)
        self.InitUI()

    def InitUI(self):

    

        # Set the frame size to match the image size
        #     self.SetSize(bitmap.GetSize())

        #   self.SetBackgroundColour(wx.BLUE)
        self.SetBackgroundColour("#D3F38F")
        menubar = wx.MenuBar()
        fileMenu = wx.Menu()
        fileItem = fileMenu.Append(wx.ID_EXIT, 'Quit', 'Quit application')
        menubar.Append(fileMenu, '&File')
        self.SetMenuBar(menubar)
        self.Bind(wx.EVT_MENU, self.OnQuit, fileItem)

        font = wx.Font(pointSize=16, family=wx.FONTFAMILY_DEFAULT, style=wx.FONTSTYLE_NORMAL,
                       weight=wx.FONTWEIGHT_NORMAL)

        vbox = wx.BoxSizer(wx.VERTICAL)

        vbox.Add((-1, 60))

        hbox1 = wx.BoxSizer(wx.HORIZONTAL)
        stock_label = wx.StaticText(self, label='Select GST Input file:')
        stock_label.SetFont(font)
        hbox1.Add(stock_label, flag=wx.LEFT | wx.RIGHT, border=10)
        vbox.Add(hbox1, flag=wx.LEFT | wx.RIGHT | wx.TOP, border=10)

        vbox.Add((-1, 20))

        hbox2 = wx.BoxSizer(wx.HORIZONTAL)
        self.file1 = wx.FilePickerCtrl(self, style=wx.FLP_USE_TEXTCTRL)
        self.file1.SetFont(font)
        hbox2.Add(self.file1, proportion=1, flag=wx.EXPAND)
        vbox.Add(hbox2, flag=wx.LEFT | wx.RIGHT | wx.TOP, border=10)

        vbox.Add((-1, 100))

        hbox7 = wx.BoxSizer(wx.HORIZONTAL)
        submit_button = wx.Button(self, label='Submit')
        hbox7.Add(submit_button, proportion=1, flag=wx.EXPAND)
        vbox.Add(hbox7, flag=wx.LEFT | wx.BOTTOM, border=10)

        self.Bind(wx.EVT_BUTTON, self.on_submit, submit_button)

        vbox.Add((-1, 20))

        font = wx.Font(10, wx.DECORATIVE, wx.ITALIC, wx.NORMAL)

        hbox8 = wx.BoxSizer(wx.HORIZONTAL)
        self.label = wx.StaticText(self, label="Welcome, RJW ELECTRICALS  - GST Grouping Program V1.0")
        self.label.SetMinSize((400, 30))  # set the minimum size of the text box
        self.label.SetFont(font)
        #     self.label.SetBackgroundColour("#FFFFFF")

        hbox8.Add(self.label, proportion=1, flag=wx.EXPAND)
        vbox.Add(hbox8, flag=wx.LEFT | wx.BOTTOM, border=10)

        # vbox.AddSpacer(100)  # Add some padding to the top

        self.gauge = wx.Gauge(self, range=100, size=(-1, 25), style=wx.GA_HORIZONTAL)
        self.gauge.SetForegroundColour(wx.Colour(0, 255, 0))  # Green color
        vbox.Add(self.gauge, proportion=0, flag=wx.EXPAND)

        self.SetSizer(vbox)

        self.update_progress(0)

        self.SetSize((700, 500))
        self.SetTitle('RJW ELECTRICALS  - GST Grouping Program V1.0')
        self.Centre()

        self.Show()

    def OnQuit(self, e):
        self.Close()

    def on_submit(self, event):

        print("Directory:")
        self.InputFile_path = self.file1.GetPath()

        self.directory = os.path.dirname(self.InputFile_path)

        self.loadInput()
        self.update_progress(20)
        self.writeoutput()

    def loadInput(self):

        self.label.SetLabel("Processing input sheet, Please wait.")
        self.Layout()

        if self.InputFile_path.endswith(".xlsx"):
            # Load the Excel file into a variable
            workbookT = openpyxl.load_workbook(self.InputFile_path)
            sheetV = workbookT.active

            self.update_progress(38)
            # Create an empty DataFrame to store the item and quantity data
            self.Inputitem_df = pd.DataFrame(
                columns=["HSN", "Description", "UQC_raw", "Total Quantity", "Total Value", "Rate", "Taxable Value",
                         "Integrated Tax Amount", "Central Tax Amount", "State/UT Tax Amount", "Cess Amount"])

            self.update_progress(40)

            # Loop through each row in the sheet
            for i, row in enumerate(sheetV.iter_rows(values_only=True)):
                # Check if both the item and quantity have a value
                if (row[0] and row[1] and row[2] and row[3] and row[4] and row[5] and row[0] != "HSN"):
                    # Add the item and quantity to the DataFrame
                    self.Inputitem_df = self.Inputitem_df.append(
                        {"HSN": row[0], "Description": row[1], "UQC_raw": row[2], "Total Quantity": row[3],
                         "Total Value": row[4], "Rate": row[5], "Taxable Value": row[6],
                         "Integrated Tax Amount": row[7], "Central Tax Amount": row[8], "State/UT Tax Amount": row[9],
                         "Cess Amount": row[10]}, ignore_index=True)


        #
        elif self.InputFile_path.endswith(".xls"):
            workbookT = xlrd.open_workbook(self.InputFile_path)
            #
            sheetV = workbookT.sheet_by_index(0)

            self.update_progress(38)
            # Create an empty DataFrame to store the item and quantity data
            self.Inputitem_df = pd.DataFrame(
                columns=["HSN", "Description", "UQC_raw", "Total Quantity", "Total Value", "Rate", "Taxable Value",
                         "Integrated Tax Amount", "Central Tax Amount", "State/UT Tax Amount", "Cess Amount"])

            self.update_progress(40)

            # Loop through each row in the sheet
            for i in range(sheetV.nrows):
                row = sheetV.row_values(i)
                # Check if both the item and quantity have a value
                #         if (row[1] and row[2] and row[1] != "Item Name") or (row[1] and row[2]== 0):
                if (row[0] and row[1] and row[2] and row[3] and row[4] and row[5] and row[0] != "HSN"):
                    # Add the item and quantity to the DataFrame
                    self.Inputitem_df = self.Inputitem_df.append(
                        {"HSN": row[0], "Description": row[1], "UQC_raw": row[2], "Total Quantity": row[3],
                         "Total Value": row[4], "Rate": row[5], "Taxable Value": row[6],
                         "Integrated Tax Amount": row[7], "Central Tax Amount": row[8], "State/UT Tax Amount": row[9],
                         "Cess Amount": row[10]}, ignore_index=True)

        else:
            raise ValueError("File format not supported. Please provide an .xls or .xlsx file.")

        self.Inputitem_df['Total Quantity'] = self.Inputitem_df['Total Quantity'].fillna(0)
        self.Inputitem_df['Total Value'] = self.Inputitem_df['Total Value'].fillna(0)
        self.Inputitem_df['Taxable Value'] = self.Inputitem_df['Taxable Value'].fillna(0)
        self.Inputitem_df['Integrated Tax Amount'] = self.Inputitem_df['Integrated Tax Amount'].fillna(0)
        self.Inputitem_df['Central Tax Amount'] = self.Inputitem_df['Central Tax Amount'].fillna(0)
        self.Inputitem_df['State/UT Tax Amount'] = self.Inputitem_df['State/UT Tax Amount'].fillna(0)
        self.Inputitem_df['Cess Amount'] = self.Inputitem_df['Cess Amount'].fillna(0)

        # replace the common text for rows where the first 4 characters match
        self.Inputitem_df['UQC'] = self.Inputitem_df['UQC_raw'].apply(
            lambda x: x[:4] + 'METERS' if x[:4] == 'MTR-' else x)

        self.Inputitem_df = self.Inputitem_df.drop(columns=['UQC_raw'])

        #    print(self.Inputitem_df)

        # remove leading and trailing spaces from the columns
        #    self.Inputitem_df['HSN'] = self.Inputitem_df['HSN'].str.strip()
        self.Inputitem_df['Description'] = self.Inputitem_df['Description'].str.strip()
        self.Inputitem_df['UQC'] = self.Inputitem_df['UQC'].str.strip()
        #   self.Inputitem_df['Rate'] = self.Inputitem_df['Rate'].str.strip()

        self.grouped_df = self.Inputitem_df.groupby(["HSN", "Description", "UQC", "Rate"]).agg(
            {"Total Quantity": 'sum', "Total Value": 'sum', "Taxable Value": 'sum', "Integrated Tax Amount": 'sum',
             "Central Tax Amount": 'sum', "State/UT Tax Amount": 'sum', "Cess Amount": 'sum'})

        # Reset the index to move the grouped columns back into regular columns
        self.grouped_df = self.grouped_df.reset_index()

        self.grouped_df['Rate'] = self.grouped_df['Rate'].apply(
            lambda x: round(float(x), 2) if isinstance(x, (int, float)) else x)
        self.grouped_df['Total Quantity'] = self.grouped_df['Total Quantity'].apply(
            lambda x: round(float(x), 2) if isinstance(x, (int, float)) else x)
        self.grouped_df['Total Value'] = self.grouped_df['Total Value'].apply(
            lambda x: round(float(x), 2) if isinstance(x, (int, float)) else x)
        self.grouped_df['Taxable Value'] = self.grouped_df['Taxable Value'].apply(
            lambda x: round(float(x), 2) if isinstance(x, (int, float)) else x)
        self.grouped_df['Integrated Tax Amount'] = self.grouped_df['Integrated Tax Amount'].apply(
            lambda x: round(float(x), 2) if isinstance(x, (int, float)) else x)
        self.grouped_df['Central Tax Amount'] = self.grouped_df['Central Tax Amount'].apply(
            lambda x: round(float(x), 2) if isinstance(x, (int, float)) else x)
        self.grouped_df['State/UT Tax Amount'] = self.grouped_df['State/UT Tax Amount'].apply(
            lambda x: round(float(x), 2) if isinstance(x, (int, float)) else x)
        self.grouped_df['Cess Amount'] = self.grouped_df['Cess Amount'].apply(
            lambda x: round(float(x), 2) if isinstance(x, (int, float)) else x)

        #     print(self.grouped_df)

    def writeoutput(self):

        self.label.SetLabel("Formatting output sheet, Please wait.")
        self.Layout()

        # Set the header colors
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')

        # Get the current date and time
        now = datetime.datetime.now()
        self.update_progress(53)

        #     directory = "D:/files_compare/out"
        self.directory = self.directory + '\Results'

        if not os.path.exists(self.directory):
            os.makedirs(self.directory)

        self.file_name = self.directory + '\Group_' + now.strftime("%Y-%m-%d_%H-%M-%S") + '.xlsx'

        # Create a Workbook object
        book = Workbook()
        self.update_progress(55)

        # Create a writer object
        writer = pd.ExcelWriter(self.file_name, engine='openpyxl')
        writer.book = book

        ################################################################################################################################
        #  Writing grouped_df sheet, formatting it and hiding it in the excel.
        #
        ###############################################################################################################################
        # Write each dataframe to a separate tab in the excel file
        self.grouped_df.to_excel(writer, index=False, float_format='%.2f', sheet_name='Grouped_Item')

        # Get a reference to the worksheet
        GroupItemSh = writer.book['Grouped_Item']

        self.update_progress(58)

        # Set the column widths
        GroupItemSh.column_dimensions['A'].width = 10
        GroupItemSh.column_dimensions['B'].width = 40
        GroupItemSh.column_dimensions['C'].width = 20
        GroupItemSh.column_dimensions['D'].width = 10
        GroupItemSh.column_dimensions['E'].width = 15
        GroupItemSh.column_dimensions['F'].width = 15
        GroupItemSh.column_dimensions['G'].width = 15
        GroupItemSh.column_dimensions['H'].width = 20
        GroupItemSh.column_dimensions['I'].width = 20
        GroupItemSh.column_dimensions['J'].width = 20
        GroupItemSh.column_dimensions['K'].width = 15

        # Change the cell format of column D
        column_letter = 'D'
        column_dimensions = GroupItemSh.column_dimensions[column_letter]
        column_dimensions.number_format = '0.00'

        # Change the cell format of column E
        column_letter = 'E'
        column_dimensions = GroupItemSh.column_dimensions[column_letter]
        column_dimensions.number_format = '0.00'

        # Change the cell format of column F
        column_letter = 'F'
        column_dimensions = GroupItemSh.column_dimensions[column_letter]
        column_dimensions.number_format = '0.00'

        GroupItemSh.column_dimensions['D'].number_format = numbers.FORMAT_NUMBER_00
        #      GroupItemSh.column_dimensions['E'].number_format = numbers.FORMAT_NUMBER_00
        #      GroupItemSh.column_dimensions['F'].number_format = numbers.FORMAT_NUMBER_00
        GroupItemSh.column_dimensions['G'].number_format = numbers.FORMAT_NUMBER_00
        GroupItemSh.column_dimensions['H'].number_format = numbers.FORMAT_NUMBER_00
        GroupItemSh.column_dimensions['I'].number_format = numbers.FORMAT_NUMBER_00
        GroupItemSh.column_dimensions['J'].number_format = numbers.FORMAT_NUMBER_00
        GroupItemSh.column_dimensions['K'].number_format = numbers.FORMAT_NUMBER_00

        for col_num, value in enumerate(self.grouped_df.columns.values):
            cell = GroupItemSh.cell(row=1, column=col_num + 1)
            cell.value = value
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        self.update_progress(60)
        # Iterate through each cell in the worksheet
        for row in GroupItemSh.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                    cell.font = Font(color='FF0000')
                    cell.border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))
                else:

                    if cell.column == 'F':
                        #      value = float(cell.value)
                        cell.value = cell.value * 1

                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))

        # Delete the sheet with the specified name
        if 'Sheet' in writer.book.sheetnames:
            writer.book.remove(writer.book['Sheet'])

        # Save the excel file
        writer.save()

        self.update_progress(100)
        self.label.SetLabel("Processing Completed.Results sheet is available in " + self.directory)

    def update_progress(self, value):
        self.gauge.SetValue(value)


def main():
    app = wx.App()
    ex = mainApp(None)
    ex.Show()
    app.MainLoop()


if __name__ == '__main__':
    main()
